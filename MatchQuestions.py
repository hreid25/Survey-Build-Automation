import openpyxl
import time
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import re

# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************

hover_words_array = []
hover_texts_array = []
for r in range(4, 100):
    hover_words = []
    hover_texts = []
    for hwc in range(4, 100, 3):
        hword = surveyhovers.cell(row=r, column=hwc).value
        hover_words.append(hword)
    for htc in range(5, 101, 3):
        htext = surveyhovers.cell(row=r, column=htc).value
        hover_texts.append(htext)
# if cell is not None and not cell.startswith('='): If a given row has hovers in some but not all languages,
# the cell is not none argument can cause the rows to not be aligned
    if hover_words != [] and hover_texts != []:
        hover_words_array.append(hover_words)
        hover_texts_array.append(hover_texts)
# hoverarray = [x for x in fhovers if x != []]

# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************

# hoverlanguage count actually includes all of the None type objects read in using
# openpyxl (approx 32), even though there may be fewer languages
# TO DO: find the last column of data and stop there (both in the question array and hover array)

hoverlanguagecount = len((hover_words_array[0]))
numtotalhovers = len(hover_words_array) - 1
numtotalquestions = len(questionarr)

langnum = 1
hoverlangnum = 0

while langnum < hoverlanguagecount:
    hovernum = 1
    while hovernum < numtotalhovers:
        word = hover_words_array[hovernum][hoverlangnum]
        text = hover_texts_array[hovernum][hoverlangnum]
        if word is not None and text is not None:
            questionnum = 1
            while questionnum < numtotalquestions:
                # + 1 is to account for the question id column
                quest = questionarr[questionnum][langnum + 1]
                if quest is not None:
                    if word.lower() in quest.lower():
                        if quest.lower().startswith(word.lower()) is True:
                            propercase = word.capitalize()
                            replacehover = "{{" + "\"" + \
                                str(propercase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum + 1] = pattern.sub(replacehover, quest)
                        else:
                            normalcase = word.lower()
                            replacehover = "{{" + "\"" + \
                                str(normalcase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum + 1] = pattern.sub(replacehover, quest)
                questionnum += 1
        hovernum += 1
    langnum += 1
    hoverlangnum += 1


# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# This section also used to determine total amount of languages in a given QIL.

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])

# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

chrome_options = Options()
chrome_options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
driver = webdriver.Chrome(chrome_options=chrome_options)

driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"


# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************

searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
# click toggle
WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
    (By.XPATH, "//*[@id='edit-tabs-dropdown']"))).click()
# Click Question Groups (Drivers) Option
WebDriverWait(driver, 20).until(EC.presence_of_element_located(
    (By.XPATH, "//*[@id='report-question-groups-nav']/a"))).click()

chgpage = 1
final_list = []
driver_match_scan = True
while driver_match_scan is True:
    # Read in Title and Slug name
    all_the_drivers = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[@id='report-question-groups-table']/tbody/tr/td[position()<3]")))
    # Create secondary list obj
    prettyname_slugname = [x.text for x in all_the_drivers]
    for counter, myvar in enumerate(prettyname_slugname):
        if myvar not in final_list:
            final_list.append(myvar)
        if counter == len(prettyname_slugname) - 1:
            try:
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='report-question-groups']/div[1]/div/ul/li[@class='next']/a"))).click()
                chgpage += 1
                # Checks to see which page is active
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]/ancestor::ul/preceding::tbody/tr")))
            except Exception:
                print("Done collecting driver 'Pretty names' and 'Slug names'!...")
                driver_match_scan = False
                # Return to page one
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@class='pagination']/li[@class='first']/a"))).click()
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()
# Create key, value pairs for pretty name and slug driver names
prettyname_slugname_dict = {final_list[i]: final_list[i + 1] for i in range(0, len(final_list), 2)}

editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()

# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************
# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************
# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************


# *************************************CUSTOM FUNCTIONS***************************************************************************
# *************************************CUSTOM FUNCTIONS************************************************************************************************************
# *************************************CUSTOM FUNCTIONS***************************************************************************

# //*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a (Next button at top of page)
#  //*[@id='survey-edit-questions']/child::div[4][@class='row']/div/div/ul/li[@class='next']/a (Next button at bottom of page)
def click_next():
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
    except Exception as c:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))))
        print("I've tried to click next, but something happened. It may be the case no Next button is present.", str(c))


def return_home():
    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)


def click_prev():
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]"))).click()


def questions_returntopageone():
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as q:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
        print(str(q))


# McLean Logo typically intercepts the click when we go to save, or click. Return to page one and occassionally when we click next
# depending on where we are on the page. It's best to click the save/next/'1'/'2' buttons at the same navbar position. (top, usually).
# selenium.common.exceptions.ElementClickInterceptedException: Message: element click intercepted:
# Element <a data-remote="true" href="/engagement/surveys/39515/edit?tab=questions">...</a> is not clickable at point
#  (368, 15). Other element would receive the click:
#  <img src="//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030" alt="Hr logo">


def save_page():
    try:
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as s:
        print(str(s))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))))


# def changelanguage():
#     try:
#         for language in emailinvitationarray[0]:
#             if language == emailinvitationarray[0][languagedropdownposition]:
#                 # WebDriverWait(driver, 20).until(
#                 #     EC.visibility_of_element_located((By.XPATH, '//*[@id="language"]')))
#                 languagedropdown = Select(driver.find_element_by_xpath('//*[@id="language"]'))
#                 languagedropdown.select_by_visible_text(language)
#     except Exception as changefail:
#         print("unable to select the language dropdown option: ", changefail)


def clickswitch():
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((
        By.XPATH, "//*[@class='tab-pane active']/div[@class='row']/div/form/child::input[@type='submit']"))).click()


def checkpagetitle():
    this_page = WebDriverWait(driver, 60).until(EC.presence_of_element_located((
        By.XPATH, "//*[@id='page-list']/div/li/fieldset[1]/h5[1]")))
    return str(this_page.text)


# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# Check for the presence of senior management relationships question groups and delete them from page 2.
# Required in order to equalize the length of the delete toggles list with that of id and text.


start_time_measure_seniormgt_deletion = time.time()
click_next()
sergeantlistdeletetoggleposition = 0
scanningforquestiongroups = True
while scanningforquestiongroups is True:
    try:
        questiongroupsobjects = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//h6[contains(text(),'Question Group Title')]")))
        findquestiongroups = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//h6[contains(text(),'Question Group Title')]")))
        seniormgmtdeletetoggles = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
            (By.XPATH, "//h6[contains(text(),'Question Group Title')]/following::div[@class='row'][3]/div/div")))
        questiongroupspresent = questiongroupsobjects.is_displayed()
        while questiongroupspresent is True:
            secondarytextlist = [x.text for x in findquestiongroups]
            for i in secondarytextlist:
                if i == "Question Group Title":
                    seniormgmtdeletetoggles[sergeantlistdeletetoggleposition].click()
                    sergeantlistdeletetoggleposition += 1
                    print(i)
            questiongroupspresent = False
            scanningforquestiongroups = False
        save_page()
        # questions_returntopageone()
        break
    except Exception:
        scanningforquestiongroups = False
        questiongroupspresent = False
        print("No question group objects found")
        # questions_returntopageone()
        break
print("--- %s seconds ---" % (time.time() - start_time_measure_seniormgt_deletion))

# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# This code block is responsible for:
# Sending text from the multilingual QIL to Sergeant Question text area fields.
# Deleting Questions (Via toggledelete())
# Adding new questions to Sergeant
# Multiple time.sleep()'s have been introduced to allow for the DOM/page to load properly. Although inefficient,
# it is difficult to improve the XPATH's for these objects as unique attribute identifiers are not used in the XML DOM.
# Also responsible for deleting questions where multilingual QIL cells are of NoneType.


# TO DO: After the code runs it should print out operations to a log file for review by the Project Coordinator for QA

# edittextarea = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
#     (By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')][@placeholder='" + (questionlistobject[2]) + "']/following::strong[position()=2 and contains(text(),'" + (questionlistobject[1]) + "')]")))

processing_qil = True
while processing_qil is True:
    deleting_questions = True
    adding_questions = True
    editing_questions = True
    while deleting_questions is True:
        numofdeletions = [x for x in questionarr if x[2] is None and x[1] is not None]
        for counter, questionlistobject in enumerate(questionarr):
            # Currently the loop will break if there are no deletions or if deletions have already been run. Issue submitted for this.
            if questionlistobject[2] is None and questionlistobject[1] is not None:
                # Click the delete toggle
                try:
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
                        (By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(questionlistobject[1]) + "')]/ancestor::div[@class='move-content sortable-disabled']/child::div[@class='row']/div/div"))).click()
                    print(counter, "Question ID: " +
                          str(questionlistobject[1]) + " was toggled Delete = ON.")
                except Exception:
                    save_page()
                    click_next()
                    print("Going to next page for Question ID:  " +
                          str(questionlistobject[1]))
                    # Check for visibility of the delete toggle associated with our question ID
                    WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(questionlistobject[1]) + "')]")))
                    # click aforementioned delete toggle if visible.
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(
                        questionlistobject[1]) + "')]/ancestor::div[@class='move-content sortable-disabled']/child::div[@class='row']/div/div"))).click()
                    continue
                # introduce finally loop to handle all other edge cases. return to page two, set deleting_questions = false
                # rotate through all pages gathering sergeantid's. then compare those to each questionlistobject[1]. If the object is not in them
                # list then skip that ID.

        save_page()
        click_prev()
        deleting_questions = False
    while adding_questions is True:
        # Collect Question driver names from page
        questiondrivernames = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//*[@id='survey_pages_attributes_0_page_questions_attributes_0_title']/following::h4[not(@*)]")))
        secondarydrivernamelist = [x.text for x in questiondrivernames]
        for i, excelrowlistobject in enumerate(questionarr):
            if excelrowlistobject[1] is None and excelrowlistobject[2] is not None:
                # OR condition to check for Y/N toggle for custom question where ID is present
                if excelrowlistobject[2] == "Empty Slot":
                    continue
                else:
                    for key, value in prettyname_slugname_dict.items():
                        if excelrowlistobject[0] == str(key):
                            slug_driver_name = value.replace("_", " ").title()
                            # Click the Add Question Button following the driver name match
                            WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
                                (By.XPATH, "//h4[not(@*) and contains(text(),'" + str(slug_driver_name) + "')]/following::button[position()=1]"))).click()
                            # Instantiate AddQuestionTextArea
                            addquestiontextarea = WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                                (By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-body']/child::div[@class='fields']/div/div/div[@class='selectize-control grouped_select optional single']/div/input")))
                            # Send question text to text area field
                            addquestiontextarea.send_keys(excelrowlistobject[2])
                            # Click Save
                            clickercounter = 1
                            while clickercounter < 2:
                                try:
                                    driver.find_element_by_tag_name('body').send_keys(Keys.TAB)
                                    WebDriverWait(driver, 1).until(                                 # Changed the driver to 1 second here. Adjust back to 2 if issues
                                        EC.element_to_be_clickable((By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-footer']/input[@type='submit']"))).click()
                                except Exception:
                                    print("Have attempted to click the save button: " +
                                          str(clickercounter) + " times.")
                                    clickercounter += 1
                                    continue
                                # Grab the Question ID of the new question and add it to questionarr[questionnum][1]
                            newcustomquestionid = WebDriverWait(driver, 10).until(
                                EC.visibility_of_element_located((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled question')][@placeholder='" + excelrowlistobject[2] + "']/following::strong[position()=2]")))
                            # Match the text from excelrowlistobject to questionarr[counter][2] (the QIL question text) and then insert the Question ID to prev column
                            for counter, qilrowlistobject in enumerate(questionarr):
                                if excelrowlistobject[2] == questionarr[counter][2]:
                                    insertcustomidtoarray = newcustomquestionid.text
                                    questionarr[counter][1] = insertcustomidtoarray
            print(questionarr[i][1])
        adding_questions = False
    while editing_questions is True:
        columnnumber = 2
        return_home()  # recently added without testing
        questions_returntopageone()
        for edit_counter, qilrowlistobj in enumerate(questionarr):
            sergeant_questionid_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                (By.XPATH, "//*[@class='switch-container span10']/div[3]/span/strong[contains(text(),'" + qilrowlistobj[1] + "']")))
            if qilrowlistobj[1] == sergeant_questionid_element:
                try:
                    sergeant_question_text_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@class='switch-container span10']/div[3]/span/strong[contains(text(),'" + qilrowlistobj[1] + "']/ancestor::div[position()=4]/textarea")))
                    QILarrayenumerated = [(index, row.index(int(qilrowlistobj[1])))
                                          for index, row in enumerate(questionarr) if int(qilrowlistobj[1]) in row]
                    question_position_rownum = QILarrayenumerated[0][0]
                    sergeant_question_text_element.clear()
                    sergeant_question_text_element.send_keys(
                        questionarr[question_position_rownum][columnnumber])
                except Exception as edit_error:
                    print("Looks like we ran out of questions to edit! Clicking Next...", edit_error)
                    save_page()
                    click_next()
                    continue
    processing_qil = False
    editing_questions = False
print("--- %s seconds ---" % (time.time() - start_time))
