import openpyxl
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

wb = openpyxl.load_workbook('QIL Document_V2_20210518_2.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']


# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])
# print(questionarr[5][2])

# for counter, questionlistobject in enumerate(questionarr):
#     print(counter)
#     if counter == len(questionarr) - 1:
#         print("ive reached the end of the index")
numofdeletions = [x for x in questionarr if x[2] is None and x[1] is not None]
print(len(numofdeletions))

# def addcustomquestions():
#     questiondrivernames = WebDriverWait(driver, 20).until(
#         EC.presence_of_all_elements_located((By.XPATH, "//*[@id='survey_pages_attributes_0_page_questions_attributes_0_title']/following::h4[not(@*)]")))
#     secondarydrivernamelist = []
#     for x in questiondrivernames:
#         secondarydrivernamelist.append(x.text)
#     for i, excelrowlistobject in enumerate(questionarr):
#         if excelrowlistobject[1] is None and excelrowlistobject[2] is not None:
#             # OR condition to check for Y/N toggle for custom question where ID is present
#             if excelrowlistobject[2] == "Empty Slot":
#                 continue
#             else:
#                 for drivername in secondarydrivernamelist:
#                     if excelrowlistobject[0] == drivername:
#                         # print(type(drivername))
#                         # Click the Add Question Button following the driver name match
#                         WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
#                             (By.XPATH, "//h4[not(@*) and contains(text(),'" + excelrowlistobject[0] + "')]/following::button[position()=1]"))).click()
#                         # Instantiate AddQuestionTextArea
#                         addquestiontextarea = WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
#                             (By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-body']/child::div[@class='fields']/div/div/div[@class='selectize-control grouped_select optional single']/div/input")))
#                         # Send question text to text area field
#                         addquestiontextarea.send_keys(excelrowlistobject[2])
#                         # Click Save
#                         clickercounter = 0
#                         while clickercounter < 2:
#                             try:
#                                 WebDriverWait(driver, 4).until(
#                                     EC.element_to_be_clickable((By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-footer']/input[@type='submit']"))).click()
#                             except Exception:
#                                 print("Have attempted: " + str(clickercounter) + " save button clicks")
#                                 clickercounter += 1
#                                 continue
#                             # Grab the Question ID of the new question and add it to questionarr[questionnum][1]
#                         newcustomquestionid = WebDriverWait(driver, 20).until(
#                             EC.visibility_of_element_located((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled question')][@placeholder='" + excelrowlistobject[2] + "']/following::strong[position()=2]")))
#                         # //textarea[contains(text(),)]/following: : strong[position() = 2]
#                         # Match the text from excelrowlistobject to questionarr[counter][2] (the QIL question text) and then insert the Question ID to prev column
#                         for counter, qilrowlistobject in enumerate(questionarr):
#                             if excelrowlistobject[2] == questionarr[counter][2]:
#                                 insertcustomidtoarray = newcustomquestionid.text
#                                 questionarr[counter][1] = insertcustomidtoarray
#         print(questionarr[i][1])


# Example below of matching the newly saved question and grabbing that new Question's ID:
# f"//*[starts-with(@class,'question-text-area sortable-disabled question')][@placeholder={excelrowlistobject[2]}]/following::strong[position()=2]"

# //*[starts-with(@class,'question-text-area sortable-disabled question')]/following::strong[position()=2]
# pass the name of the driver back to the web element object to be found and clicked
# f//h4[not(@*)contains(text()='{excelrowlistobject[0]}')]
# f"//h4[not(@*) and contains(text(),{excelrowlistobject[0]})]/following::button[position()=1]")))

# need to return the match position in the QIL
# print(addquestion[2])
# print(questionstobeadded)
# print(drivername)

# print(questionarr[i][1])

# list = [i for x in questionarr[i][rownum] if x is not None]
# print(list)
# def addquestions():
#     if questionarr[questionnum][1] is None and questionarr[questionnum][languagenumber] is not None:
#         drivername = [x for x in questionarr[questionnum][0]]
#         addquestionlist = [i for question in questionarr if question[0] is not None]
