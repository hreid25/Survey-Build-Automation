import openpyxl

wb = openpyxl.load_workbook('QIL Document_V2_20210518_2.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']
# ,
qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring
    # print(i, category[1])

print(questionarr[5][2])
