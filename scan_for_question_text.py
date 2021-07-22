import openpyxl
import time
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import re
from webdriver_manager.chrome import ChromeDriverManager

# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************

hover_words_array = []
hover_texts_array = []
for r in range(4, 100):
    hover_words = []
    hover_texts = []
    for hwc in range(4, 100, 3):
        hword = surveyhovers.cell(row=r, column=hwc).value
        hover_words.append(hword)
    for htc in range(5, 101, 3):
        htext = surveyhovers.cell(row=r, column=htc).value
        hover_texts.append(htext)
# if cell is not None and not cell.startswith('='): If a given row has hovers in some but not all languages,
# the cell is not none argument can cause the rows to not be aligned
    if hover_words != [] and hover_texts != []:
        hover_words_array.append(hover_words)
        hover_texts_array.append(hover_texts)
# hoverarray = [x for x in fhovers if x != []]

# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************

# hoverlanguage count actually includes all of the None type objects read in using
# openpyxl (approx 32), even though there may be fewer languages
# TO DO: find the last column of data and stop there (both in the question array and hover array)

hoverlanguagecount = len((hover_words_array[0]))
numtotalhovers = len(hover_words_array) - 1
numtotalquestions = len(questionarr)

langnum = 1
hoverlangnum = 0

while langnum < hoverlanguagecount:
    hovernum = 1
    while hovernum < numtotalhovers:
        word = hover_words_array[hovernum][hoverlangnum]
        text = hover_texts_array[hovernum][hoverlangnum]
        if word is not None and text is not None:
            questionnum = 1
            while questionnum < numtotalquestions:
                # + 1 is to account for the question id column
                quest = questionarr[questionnum][langnum + 1]
                if quest is not None:
                    if word.lower() in quest.lower():
                        if quest.lower().startswith(word.lower()) is True:
                            propercase = word.capitalize()
                            replacehover = "{{" + "\"" + \
                                str(propercase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum + 1] = pattern.sub(replacehover, quest)
                        else:
                            normalcase = word.lower()
                            replacehover = "{{" + "\"" + \
                                str(normalcase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum + 1] = pattern.sub(replacehover, quest)
                questionnum += 1
        hovernum += 1
    langnum += 1
    hoverlangnum += 1


# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# This section also used to determine total amount of languages in a given QIL.

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])

# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

options = Options()
options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
# driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"

# *************************************CUSTOM FUNCTIONS***************************************************************************
# *************************************CUSTOM FUNCTIONS************************************************************************************************************
# *************************************CUSTOM FUNCTIONS***************************************************************************

# //*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a (Next button at top of page)
#  //*[@id='survey-edit-questions']/child::div[4][@class='row']/div/div/ul/li[@class='next']/a (Next button at bottom of page)

# McLean Logo typically intercepts the click when we go to save, or click. Return to page one and occassionally when we click next
# depending on where we are on the page. It's best to click the save/next/'1'/'2' buttons at the same navbar position. (top, usually).
# selenium.common.exceptions.ElementClickInterceptedException: Message: element click intercepted:
# Element <a data-remote="true" href="/engagement/surveys/39515/edit?tab=questions">...</a> is not clickable at point
#  (368, 15). Other element would receive the click:
#  <img src="//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030" alt="Hr logo">


def click_next():
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
    except Exception as c:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))))
        print("I've tried to click next, but something happened. It may be the case no Next button is present.", str(c))


def return_home():
    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)


def click_prev():
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]"))).click()


def questions_returntopageone():
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as q:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
        print(str(q))


def save_page():
    try:
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as s:
        print(str(s))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))))


def click_first_questions_page():
    try:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except Exception as click_first_error:
        print(click_first_error)

# def changelanguage():
#     try:
#         for language in emailinvitationarray[0]:
#             if language == emailinvitationarray[0][languagedropdownposition]:
#                 # WebDriverWait(driver, 20).until(
#                 #     EC.visibility_of_element_located((By.XPATH, '//*[@id="language"]')))
#                 languagedropdown = Select(driver.find_element_by_xpath('//*[@id="language"]'))
#                 languagedropdown.select_by_visible_text(language)
#     except Exception as changefail:
#         print("unable to select the language dropdown option: ", changefail)


def clickswitch():
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((
        By.XPATH, "//*[@class='tab-pane active']/div[@class='row']/div/form/child::input[@type='submit']"))).click()


def checkpagetitle():
    this_page = WebDriverWait(driver, 60).until(EC.presence_of_element_located((
        By.XPATH, "//*[@id='page-list']/div/li/fieldset[1]/h5[1]")))
    return str(this_page.text)

# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# Go into Question Groups and collect all driver slugs and custom names (pretty names)


searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
# click toggle
WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
    (By.XPATH, "//*[@id='edit-tabs-dropdown']"))).click()
# Click Question Groups (Drivers) Option
WebDriverWait(driver, 20).until(EC.presence_of_element_located(
    (By.XPATH, "//*[@id='report-question-groups-nav']/a"))).click()

chgpage = 1
final_list = []
driver_match_scan = True
while driver_match_scan is True:
    # Read in Title and Slug name
    all_the_drivers = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[@id='report-question-groups-table']/tbody/tr/td[position()<3]")))
    # Create secondary list obj
    prettyname_slugname = [x.text for x in all_the_drivers]
    for counter, myvar in enumerate(prettyname_slugname):
        if myvar not in final_list:
            final_list.append(myvar)
        if counter == len(prettyname_slugname) - 1:
            try:
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='report-question-groups']/div[1]/div/ul/li[@class='next']/a"))).click()
                chgpage += 1
                # Checks to see which page is active
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]/ancestor::ul/preceding::tbody/tr")))
            except Exception:
                print("Done collecting driver 'Pretty names' and 'Slug names'!...")
                driver_match_scan = False
                # Return to page one
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@class='pagination']/li[@class='first']/a"))).click()
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()
# Create key, value pairs for pretty name and slug driver names
prettyname_slugname_dict = {final_list[i]: final_list[i + 1] for i in range(0, len(final_list), 2)}

editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()


# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************
# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************
# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************

current_sergeant_question_text_list = []
current_sergeant_question_id_list = []
chgpage = 1
id_match_scan = True
while id_match_scan is True:
    # Read in id text
    sergeant_question_id_list = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong")))
    sergeant_question_text_list = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[contains(@class,'question-text-area sortable-disabled') and contains(@id,'survey_pages_attributes')]")))
    # Create secondary list obj
    all_my_sergeant_ids = [x.text for x in sergeant_question_id_list]
    all_my_sergeant_q_text = [y.text for y in sergeant_question_text_list]
    for counter, myid in enumerate(all_my_sergeant_ids):
        if str(myid) not in current_sergeant_question_id_list:
            current_sergeant_question_id_list.append(str(myid))
        elif counter == len(all_my_sergeant_ids) - 1:
            try:
                chgpage += 1
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
                # Wait for detection next page being active
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]")))
            except Exception:
                print("Done retrieving all initial Sergeant ID's.")
                # Return to first page
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
                id_match_scan = False
                break
    for counter, mytxt in enumerate(all_my_sergeant_q_text):
        if mytxt not in sergeant_question_text_list:
            current_sergeant_question_text_list.append(mytxt)

scanned_sergeant_question_text = list(set(current_sergeant_question_text_list))

# for x in current_sergeant_question_text_list:
#     print(x)
# for y in current_sergeant_question_id_list:
#     print(y)
