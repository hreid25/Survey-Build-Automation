import openpyxl
import time
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import re
from webdriver_manager.chrome import ChromeDriverManager

# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************

hover_words_array = []
hover_texts_array = []
for r in range(4, 100):
    hover_words = []
    hover_texts = []
    for hwc in range(4, 100, 3):
        hword = surveyhovers.cell(row=r, column=hwc).value
        hover_words.append(hword)
    for htc in range(5, 101, 3):
        htext = surveyhovers.cell(row=r, column=htc).value
        hover_texts.append(htext)
# if cell is not None and not cell.startswith('='): If a given row has hovers in some but not all languages,
# the cell is not none argument can cause the rows to not be aligned
    if hover_words != [] and hover_texts != []:
        hover_words_array.append(hover_words)
        hover_texts_array.append(hover_texts)
# hoverarray = [x for x in fhovers if x != []]
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# hoverlanguage count actually includes all of the None type objects read in using
# openpyxl (approx 32), even though there may be fewer languages
# TO DO: find the last column of data and stop there (both in the question array and hover array)
hoverlanguagecount = len((hover_words_array[0]))
numtotalhovers = len(hover_words_array) - 1
numtotalquestions = len(questionarr)
langnum = 1
hoverlangnum = 0
while langnum < hoverlanguagecount:
    hovernum = 1
    while hovernum < numtotalhovers:
        word = hover_words_array[hovernum][hoverlangnum]
        text = hover_texts_array[hovernum][hoverlangnum]
        if word is not None and text is not None:
            questionnum = 1
            while questionnum < numtotalquestions:
                # + 1 is to account for the question id column
                quest = questionarr[questionnum][langnum + 1]
                if quest is not None:
                    if word.lower() in quest.lower():
                        if quest.lower().startswith(word.lower()) is True:
                            propercase = word.capitalize()
                            replacehover = "{{" + "\"" + \
                                str(propercase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum +
                                                     1] = pattern.sub(replacehover, quest)
                        else:
                            normalcase = word.lower()
                            replacehover = "{{" + "\"" + \
                                str(normalcase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum +
                                                     1] = pattern.sub(replacehover, quest)
                questionnum += 1
        hovernum += 1
    langnum += 1
    hoverlangnum += 1

# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# This section also used to determine total amount of languages in a given QIL.

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])

# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

options = Options()
options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
# driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"

# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************
# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************
# ***********************FIND SURVEY NAME / START EDITING QUESTIONS********************************************************


searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
# click toggle
WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
    (By.XPATH, "//*[@id='edit-tabs-dropdown']"))).click()
# Click Question Groups (Drivers) Option
WebDriverWait(driver, 20).until(EC.presence_of_element_located(
    (By.XPATH, "//*[@id='report-question-groups-nav']/a"))).click()

# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
chgpage = 1
final_list = []
driver_match_scan = True
while driver_match_scan is True:
    # Read in Title and Slug name
    all_the_drivers = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[@id='report-question-groups-table']/tbody/tr/td[position()<3]")))
    # Create secondary list obj
    prettyname_slugname = [x.text for x in all_the_drivers]
    for counter, myvar in enumerate(prettyname_slugname):
        if myvar not in final_list:
            final_list.append(myvar)
        if counter == len(prettyname_slugname) - 1:
            try:
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='report-question-groups']/div[1]/div/ul/li[@class='next']/a"))).click()
                chgpage += 1
                # Checks to see which page is active
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]/ancestor::ul/preceding::tbody/tr")))
            except Exception:
                print("Done collecting Driver 'Pretty names' and 'Slug names'!...")
                driver_match_scan = False
                # Return to page one
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@class='pagination']/li[@class='first']/a"))).click()
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()
# Create key, value pairs for pretty name and slug driver names
prettyname_slugname_dict = {final_list[i]: final_list[i + 1] for i in range(0, len(final_list), 2)}


#  CLICK EDIT QUESTIONS BUTTON
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()


# *************************************CUSTOM FUNCTIONS***************************************************************************
# *************************************CUSTOM FUNCTIONS************************************************************************************************************
# *************************************CUSTOM FUNCTIONS***************************************************************************

# //*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a (Next button at top of page)
#  //*[@id='survey-edit-questions']/child::div[4][@class='row']/div/div/ul/li[@class='next']/a (Next button at bottom of page)
def click_next():
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
        # WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
        #     (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='" + str(int(pagenum + 1)) + "']")))
    except Exception as c:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))))
        print("I've tried to click next, but no Next button is present.", str(c))


def return_home():
    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)


def click_prev():
    # WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
    #     (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]"))).click()
    active_page = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]")))
    # implement check here to ensure we are going to the aforementioned page.
    try:
        prev_page_num = int(active_page.text) - 1
        active_page.click()
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'" + str(prev_page_num) + "')]")))
    except Exception as prev_click_error:
        print('Something happened when click_prev() was invoked', prev_click_error)


def questions_returntopageone():
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as q:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
        print(str(q))


def save_page():
    try:
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as s:
        print(str(s))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))))


def click_first_questions_page():
    try:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except Exception as click_first_error:
        print(click_first_error)


def escape_string_for_xpath(s):
    if '"' in s and "'" in s:
        return 'concat(%s)' % ", '\"',".join('"%s"' % x for x in s.split('"'))
    elif '"' in s:
        return "'%s'" % s
    return '"%s"' % s

    # USAGE EXAMPLE
    # escaped_title = escape_string_for_xpath('"that\'ll be the "day"')
    # driver.find_element_by_xpath('//a[@title=' + escaped_title + ']')
    # Note that it adds the appropriate kind of quotes around your string, so be sure not to add extra quotes around the return value.
    # https://newbedev.com/how-to-escape-single-quote-in-xpath-1-0-in-selenium-for-python?fbclid=IwAR1z51gplypVxKsFC_gx76JBlfIw5YTCBMwhhQikXB0BP6fNc8s3OR8x_YM


# def changelanguage():
#     try:
#         for language in emailinvitationarray[0]:
#             if language == emailinvitationarray[0][languagedropdownposition]:
#                 # WebDriverWait(driver, 20).until(
#                 #     EC.visibility_of_element_located((By.XPATH, '//*[@id="language"]')))
#                 languagedropdown = Select(driver.find_element_by_xpath('//*[@id="language"]'))
#                 languagedropdown.select_by_visible_text(language)
#     except Exception as changefail:
#         print("unable to select the language dropdown option: ", changefail)


def clickswitch():
    WebDriverWait(driver, 60).until(EC.visibility_of_element_located((
        By.XPATH, "//*[@class='tab-pane active']/div[@class='row']/div/form/child::input[@type='submit']"))).click()


def checkpagetitle():
    this_page = WebDriverWait(driver, 60).until(EC.presence_of_element_located((
        By.XPATH, "//*[@id='page-list']/div/li/fieldset[1]/h5[1]")))
    return str(this_page.text)

# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************
# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************
# ***********************GATHER SERGEANT IDS BEFORE EDITING, DELETING OR ADDING*****************************************************************************************


current_sergeant_question_text_list = []
current_sergeant_question_id_list = []
chgpage = 1
id_match_scan = True
while id_match_scan is True:
    # Read in id text
    sergeant_question_id_list = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong")))
    sergeant_question_text_list = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[contains(@class,'question-text-area sortable-disabled') and contains(@id,'survey_pages_attributes')]")))
    # Create secondary list obj
    all_my_sergeant_ids = [x.text for x in sergeant_question_id_list]
    all_my_sergeant_q_text = [y.text for y in sergeant_question_text_list]
    for counter, myid in enumerate(all_my_sergeant_ids):
        if str(myid) not in current_sergeant_question_id_list:
            current_sergeant_question_id_list.append(str(myid))
        elif counter == len(all_my_sergeant_ids) - 1:
            try:
                chgpage += 1
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
                # Wait for detection next page being active
                WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]")))
            except Exception:
                print("Done retrieving all initial Sergeant ID's.")
                # Return to first page
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
                id_match_scan = False
                break
    for counter, mytxt in enumerate(all_my_sergeant_q_text):
        if mytxt not in sergeant_question_text_list:
            current_sergeant_question_text_list.append(mytxt)

scanned_sergeant_question_text = list(set(current_sergeant_question_text_list))

# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# ***********************REMOVE SENIOR MANAGEMENT QUESTIONS GROUPING*****************************************************************************************
# Check for the presence of senior management relationships question groups and delete them from page 2.
# Required in order to equalize the length of the delete toggles list with that of id and text.

start_time_measure_seniormgt_deletion = time.time()
click_next()
sergeantlistdeletetoggleposition = 0
scanningforquestiongroups = True
while scanningforquestiongroups is True:
    try:
        questiongroupsobjects = WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
            (By.XPATH, "//h6[contains(text(),'Question Group Title')]")))
        findquestiongroups = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.XPATH, "//h6[contains(text(),'Question Group Title')]")))
        seniormgmtdeletetoggles = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
            (By.XPATH, "//h6[contains(text(),'Question Group Title')]/following::div[@class='row'][3]/div/div")))
        questiongroupspresent = questiongroupsobjects.is_displayed()
        while questiongroupspresent is True:
            secondarytextlist = [str(x.text) for x in findquestiongroups]
            for i in secondarytextlist:
                if i == "Question Group Title":
                    seniormgmtdeletetoggles[sergeantlistdeletetoggleposition].click()
                    sergeantlistdeletetoggleposition += 1
                    # print(i)
            questiongroupspresent = False
            scanningforquestiongroups = False
        save_page()
        # questions_returntopageone()
        break
    except Exception:
        scanningforquestiongroups = False
        questiongroupspresent = False
        print("Done scanning for senior management question grouping + deleting...")
        # questions_returntopageone()
        break
print("--- %s seconds ---" % (time.time() - start_time_measure_seniormgt_deletion))

# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# ***********************COMPARE QUESTION IDS IN SERGEANT TO MULTI-QIL*****************************************************************************************
# This code block is responsible for:
# Sending text from the multilingual QIL to Sergeant Question text area fields.
# Deleting Questions (Via toggledelete())
# Adding new questions to Sergeant
# Also responsible for deleting questions where multilingual QIL cells are of NoneType.
# TO DO: After the code runs it should print out operations to a log file for review by the Project Coordinator for QA

pagenum = 1
processing_qil = True
while processing_qil is True:
    deleting_questions = True
    adding_questions = True
    editing_questions = True
    while deleting_questions is True:
        for counter, questionlistobject in enumerate(questionarr):
            if str(questionlistobject[1]) not in current_sergeant_question_id_list:
                print("Question ID: ", questionlistobject[1], "has already been deleted")
                continue
            # Another if condition in here to check the toggle status (Y/N)
            elif questionlistobject[2] is None and questionlistobject[1] is not None and str(questionlistobject[1]) in current_sergeant_question_id_list:
                try:
                    # Click delete
                    WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
                        (By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(questionlistobject[1]) + "')]/ancestor::div[@class='move-content sortable-disabled']/child::div[@class='row']/div/div"))).click()
                    print(counter, "Question ID: " +
                          str(questionlistobject[1]) + " was toggled Delete = ON.")
                except Exception as my_error:
                    if pagenum < 3:
                        save_page()
                        print("Going to next page for Question ID:  " +
                              str(questionlistobject[1]))
                        pagenum += 1
                        click_next()
                        # check for visibility of the delete toggle associated with our question ID
                        WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
                            (By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(questionlistobject[1]) + "')]")))
                        # click aforementioned delete toggle if visible.
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled')]/following::strong[position()=2 and contains(text(),'" + str(
                            questionlistobject[1]) + "')]/ancestor::div[@class='move-content sortable-disabled']/child::div[@class='row']/div/div"))).click()
                        continue
                    else:
                        print(
                            "Some other condition has been triggered when trying to delete questions: ", my_error)
                        deleting_questions = False
                        break
            elif counter == len(questionarr) - 1:
                print('===========================Done deleting Questions!===========================')
                save_page()
                # Return to top of page (ctrl + home). Otherwise, clicking save fails
                return_home()
                # Go To Page 1
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
                WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@class='page active']/a[text()='1']")))
                deleting_questions = False
                pagenum = 1
                # processing_qil = False
                break
    click_next()
    while adding_questions is True:
        # Still does not account for adding questions on page 3 (Open Text Questions)
        for i, excelrowlistobject in enumerate(questionarr):
            # Check if ID is None and Cell text is not None.
            if excelrowlistobject[1] is None and excelrowlistobject[2] is not None:
                # require an OR condition to check for Y/N toggle for custom question where ID is present -- handle like a new question
                # All questions overwrite a default cell in the QIL that says 'Empty Slot'
                if excelrowlistobject[2] == "Empty Slot":
                    continue
                elif str(excelrowlistobject[2]) in scanned_sergeant_question_text:
                    # Pass the text through to the xpath and then grab the custom question ID, reinsert to QIL array.
                    # Need to format the string to be escaped when its passed to the XPATH expression to account for apostrophe's and double quotes that would otherwise act as line terminators.
                    # https://newbedev.com/how-to-escape-single-quote-in-xpath-1-0-in-selenium-for-python?fbclid=IwAR1z51gplypVxKsFC_gx76JBlfIw5YTCBMwhhQikXB0BP6fNc8s3OR8x_YM
                    escaped_question_text = escape_string_for_xpath(excelrowlistobject[2])
                    newcustomquestionid = WebDriverWait(driver, 20).until(
                        EC.visibility_of_element_located((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled question')][@placeholder=" + escaped_question_text + "]/following::strong[position()=2]")))
                    insertcustomidtoarray = int(newcustomquestionid.text)
                    questionarr[i][1] = insertcustomidtoarray
                    print('Question ID: ' +
                          str(excelrowlistobject[1]) + " has already been added to the survey.")
                    continue
                # Match the driver name to its slug and return that value to our xpath below
                for key, value in prettyname_slugname_dict.items():
                    if excelrowlistobject[0] == str(key):
                        slug_driver_name = value.replace("_", " ").title()
                if str(excelrowlistobject[2]) not in scanned_sergeant_question_text:
                    # Click add question btn
                    # ISSUES: on the last loop the newcustomquestionid assignment fails when calling webdriver wait.
                    WebDriverWait(driver, 5).until(EC.invisibility_of_element(
                        (By.XPATH, "/html/body/div[2][@class='modal-backdrop fade in']")))
                    WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
                        (By.XPATH, "//h4[not(@*) and contains(text(),'" + str(slug_driver_name) + "')]/following::button[position()=1]"))).click()
                    # Instantiate AddQuestionTextArea
                    addquestiontextarea = WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                        (By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-body']/child::div[@class='fields']/div/div/div[@class='selectize-control grouped_select optional single']/div/input")))
                    # Send question text to text area field
                    addquestiontextarea.send_keys(excelrowlistobject[2])
                    clickercounter = 1
                    while clickercounter < 2:
                        try:
                            # Return to top of page and Click Save
                            driver.find_element_by_tag_name('body').send_keys(Keys.TAB)
                            WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, "//form[@id='add-custom-question-form']/div[@class='modal-footer']/input[@type='submit']"))).click()
                            # check to see visibility of new question added, grab question id and reinsert to question array.
                            escaped_question_text = escape_string_for_xpath(excelrowlistobject[2])
                            newcustomquestionid = WebDriverWait(driver, 60).until(
                                EC.visibility_of_element_located((By.XPATH, "//*[starts-with(@class,'question-text-area sortable-disabled question')][@placeholder=" + escaped_question_text + "]/following::strong[position()=2]")))
                            insertcustomidtoarray = int(newcustomquestionid.text)
                            questionarr[i][1] = insertcustomidtoarray
                        except Exception:
                            clickercounter += 1
                            continue
                    print('======================================================')
                    print('slug driver name: ', slug_driver_name.upper())
                    print('Question Text: ', excelrowlistobject[2])
                    print('Question ID: ', questionarr[i][1])
                    print('======================================================')
        print("===========================Done adding questions===========================")
        adding_questions = False
        # processing_qil = False
        questions_returntopageone()
        # wait for page 1 to be visible
        WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='1']")))
    editing_questions = True
    while editing_questions is True:
        columnnumber = 2
        # implement a wait to ensure we are on page 1
        questionareaboxeslist = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
            (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong/ancestor::div/textarea[starts-with(@class,'question-text-area sortable-disabled question')]")))
        sergeantquestionidlist = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
            (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong")))
        secondaryidlist = [x.text for x in sergeantquestionidlist]
        # print(secondaryidlist)
        for idposcounter, myquestionid in enumerate(secondaryidlist):
            QILarrayenumerated = [(index, row.index(int(myquestionid)))
                                  for index, row in enumerate(questionarr) if int(myquestionid) in row]
            print(QILarrayenumerated, "Question ID: ", myquestionid)
            QILquestionposition = QILarrayenumerated[0][0]
            # print(questionarr[QILquestionposition][columnnumber])
            if questionarr[QILquestionposition][columnnumber] is not None:
                replacetext = questionarr[QILquestionposition][columnnumber]
                questionareaboxeslist[idposcounter].clear()
                questionareaboxeslist[idposcounter].send_keys(replacetext)
                if int(secondaryidlist[0]) == int(16595):
                    save_page()
                    # wait for visibility of Successful save banner
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@class='alert alert-notice']/a")))
                    click_next()
                    # wait for page 2 to be the active page
                    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='2']")))
                    pagenum = 2
                # This checks if we have edited the last question of page 2, then saves and clicks next
                elif idposcounter == len(secondaryidlist) - 1 and pagenum == 2:
                    save_page()
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@class='alert alert-notice']/a")))
                    click_next()
                    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='3']")))
                    pagenum = 3
                # This checks to see if we finished editing questions on page (3), saves and returns to first
                elif idposcounter == len(secondaryidlist) - 1 and pagenum == 3:
                    save_page()
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@class='alert alert-notice']/a")))
                    questions_returntopageone()
                    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='1']")))
                    editing_questions = False
                    processing_qil = False
                    break
    print("--- %s seconds ---" % (time.time() - start_time))
