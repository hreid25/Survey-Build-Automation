import openpyxl
import time
# import logging
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import Select
import re


# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************

hover_words_array = []
hover_texts_array = []
for r in range(4, 100):
    hover_words = []
    hover_texts = []
    for hwc in range(4, 100, 3):
        hword = surveyhovers.cell(row=r, column=hwc).value
        hover_words.append(hword)
    for htc in range(5, 101, 3):
        htext = surveyhovers.cell(row=r, column=htc).value
        hover_texts.append(htext)
# if cell is not None and not cell.startswith('='): If a given row has hovers in some but not all languages,
# the cell is not none argument can cause the rows to not be aligned
    if hover_words != [] and hover_texts != []:
        hover_words_array.append(hover_words)
        hover_texts_array.append(hover_texts)
# hoverarray = [x for x in fhovers if x != []]
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# hoverlanguage count actually includes all of the None type objects read in using
# openpyxl (approx 32), even though there may be fewer languages
# TO DO: find the last column of data and stop there (both in the question array and hover array)
# hoverlanguagecount = len((hover_words_array[0]))
# numtotalhovers = len(hover_words_array) - 1
# numtotalquestions = len(questionarr)
# langnum = 1
# hoverlangnum = 0
# while langnum < hoverlanguagecount:
#     hovernum = 1
#     while hovernum < numtotalhovers:
#         word = hover_words_array[hovernum][hoverlangnum]
#         text = hover_texts_array[hovernum][hoverlangnum]
#         if word is not None and text is not None:
#             questionnum = 1
#             while questionnum < numtotalquestions:
#                 # + 1 is to account for the question id column
#                 quest = questionarr[questionnum][langnum + 1]
#                 if quest is not None:
#                     if word.lower() in quest.lower():
#                         if quest.lower().startswith(word.lower()) is True:
#                             propercase = word.capitalize()
#                             replacehover = "{{" + "\"" + \
#                                 str(propercase) + " (" + str(text) + ")\" |hover}} "
#                             pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
#                             questionarr[questionnum][langnum +
#                                                      1] = pattern.sub(replacehover, quest)
#                         else:
#                             normalcase = word.lower()
#                             replacehover = "{{" + "\"" + \
#                                 str(normalcase) + " (" + str(text) + ")\" |hover}} "
#                             pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
#                             questionarr[questionnum][langnum +
#                                                      1] = pattern.sub(replacehover, quest)
#                 questionnum += 1
#         hovernum += 1
#     langnum += 1
#     hoverlangnum += 1

# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# This section also used to determine total amount of languages in a given QIL.

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])

# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************


def xpath_string_escape(input_str):
    # """ creates a concatenation of alternately-quoted strings that is always a valid XPath expression """
    parts = input_str.split("'")
    return "concat('" + "', \"'\" , '".join(parts) + "', '')"
    newstring = xpath_string_escape(excelrowlistobject[2])
    print(newstring)


for i, excelrowlistobject in enumerate(questionarr):
    xpath_string_escape(excelrowlistobject[2])
    # apos = "'"
    # if apos in excelrowlistobject[2]:
    #     newstring = xpath_string_escape(excelrowlistobject[2])
    #     print(newstring)
