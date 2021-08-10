import openpyxl
import time
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
# from selenium.webdriver.common.action_chains import ActionChains
# from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import re
from webdriver_manager.chrome import ChromeDriverManager

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']


def click_next():
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
        # WebDriverWait(driver, 10).until(EC.visibility_of_element_located(
        #     (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='" + str(int(pagenum + 1)) + "']")))
    except Exception as c:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))))
        print("I've tried to click next, but no Next button is present.", str(c))


def return_home():
    driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)


def click_prev():
    # WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
    #     (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]"))).click()
    active_page = WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
        (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'2')]")))
    # implement check here to ensure we are going to the aforementioned page.
    try:
        prev_page_num = int(active_page.text) - 1
        active_page.click()
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li/a[contains(text(),'" + str(prev_page_num) + "')]")))
    except Exception as prev_click_error:
        print('Something happened when click_prev() was invoked', prev_click_error)


def questions_returntopageone():
    try:
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as q:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
        print(str(q))


def save_page():
    try:
        WebDriverWait(driver, 60).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))).click()
    except ElementClickInterceptedException or StaleElementReferenceException or NoSuchElementException as s:
        print(str(s))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[5]/input[@type='submit']"))))


def click_first_questions_page():
    try:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
    except Exception as click_first_error:
        print(click_first_error)
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************


qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring

# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************
# ***********************CREATE HOVER ARRAY*****************************************************************************************

hover_words_array = []
hover_texts_array = []
for r in range(4, 100):
    hover_words = []
    hover_texts = []
    for hwc in range(4, 100, 3):
        hword = surveyhovers.cell(row=r, column=hwc).value
        hover_words.append(hword)
    for htc in range(5, 101, 3):
        htext = surveyhovers.cell(row=r, column=htc).value
        hover_texts.append(htext)
# if cell is not None and not cell.startswith('='): If a given row has hovers in some but not all languages,
# the cell is not none argument can cause the rows to not be aligned
    if hover_words != [] and hover_texts != []:
        hover_words_array.append(hover_words)
        hover_texts_array.append(hover_texts)
# hoverarray = [x for x in fhovers if x != []]
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# ***********************REPLACE WORD WITH HOVER*******************************************************************************
# hoverlanguage count actually includes all of the None type objects read in using
# openpyxl (approx 32), even though there may be fewer languages
# TO DO: find the last column of data and stop there (both in the question array and hover array)
hoverlanguagecount = len((hover_words_array[0]))
numtotalhovers = len(hover_words_array) - 1
numtotalquestions = len(questionarr)
langnum = 1
hoverlangnum = 0
while langnum < hoverlanguagecount:
    hovernum = 1
    while hovernum < numtotalhovers:
        word = hover_words_array[hovernum][hoverlangnum]
        text = hover_texts_array[hovernum][hoverlangnum]
        if word is not None and text is not None:
            questionnum = 1
            while questionnum < numtotalquestions:
                # + 1 is to account for the question id column
                quest = questionarr[questionnum][langnum + 1]
                if quest is not None:
                    if word.lower() in quest.lower():
                        if quest.lower().startswith(word.lower()) is True:
                            propercase = word.capitalize()
                            replacehover = "{{" + "\"" + \
                                str(propercase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum +
                                                     1] = pattern.sub(replacehover, quest)
                        else:
                            normalcase = word.lower()
                            replacehover = "{{" + "\"" + \
                                str(normalcase) + " (" + str(text) + ")\" |hover}} "
                            pattern = re.compile('\\b' + word + '\\s', re.IGNORECASE)
                            questionarr[questionnum][langnum +
                                                     1] = pattern.sub(replacehover, quest)
                questionnum += 1
        hovernum += 1
    langnum += 1
    hoverlangnum += 1

# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# *************************READ IN SURVEY INVITATION EMAIL**************************************************************
# This section also used to determine total amount of languages in a given QIL.

arr = []
for r in range(16, 37):
    words = []
    for c in range(3, 100, 2):
        if surveyinv.cell(row=r, column=c).value is not None:
            words.append(surveyinv.cell(row=r, column=c).value)
    arr.append(words)
emailinvitationarray = [x for x in arr if x != []]
totallanguagecount = len(emailinvitationarray[0])

# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

options = Options()
options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
# driver = webdriver.Chrome(chrome_options=chrome_options)
driver = webdriver.Chrome(executable_path=ChromeDriverManager().install(), options=options)
driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"
searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()

for i in questionarr[2][0]:
    print(i)

editing_questions = True
while editing_questions is True:
    columnnumber = 2
    # return_home()  # recently added without testing
    # questions_returntopageone()
    # implement a wait to ensure we are on page 1
    questionareaboxeslist = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong/ancestor::div/textarea[starts-with(@class,'question-text-area sortable-disabled question')]")))
    sergeantquestionidlist = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong")))
    secondaryidlist = [x.text for x in sergeantquestionidlist]
    for idposcounter, myquestionid in enumerate(secondaryidlist):
        QILarrayenumerated = [(index, row.index(int(myquestionid)))
                              for index, row in enumerate(questionarr) if int(myquestionid) in row]
        QILquestionposition = QILarrayenumerated[0][0]
        if questionarr[QILquestionposition][columnnumber] is not None:
            replacetext = questionarr[QILquestionposition][columnnumber]
            questionareaboxeslist[idposcounter].clear()
            questionareaboxeslist[idposcounter].send_keys(replacetext)
            print(QILarrayenumerated[0][0])
            if int(secondaryidlist[0]) == int(16595):
                save_page()
                # wait for visibility of Successful save banner
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@class='alert alert-notice']/a")))
                click_next()
                # wait for page 2 to be the active page
                WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='2']")))
                pagenum = 2
            # This checks if we have edited the last question of page 2, then saves and clicks next
            elif idposcounter == len(sergeantquestionidlist) and pagenum == 2:
                save_page()
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@class='alert alert-notice']/a")))
                click_next()
                WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='3']")))
                pagenum = 3
            # This checks to see if we finished editing questions on page (3), saves and returns to first
            elif idposcounter == len(sergeantquestionidlist) and pagenum == 3:
                save_page()
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@class='alert alert-notice']/a")))
                questions_returntopageone()
                WebDriverWait(driver, 20).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[@class='page active']/a[text()='1']")))
                editing_questions = False
                break
