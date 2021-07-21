import openpyxl
import time
# import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
# import re
# from selenium.webdriver.common.by import By

# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************


def click_next():
    try:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
            (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
    except Exception as c:
        WebDriverWait(driver, 20).until(EC.invisibility_of_element(
            (By.XPATH, "//sergeant-uploads1.s3.amazonaws.com/sergeant/brands/production/2/hr-logo.svg?1499654030")))
        driver.execute_script("arguments[0].click();", WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))))
        print("I've tried to click next, but something happened. It may be the case no Next button is present.", str(c))


qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring


# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

chrome_options = Options()
chrome_options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
driver = webdriver.Chrome(chrome_options=chrome_options)

driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"

# **********************************scan sergeant for existing id positions********************************************************
# **********************************scan sergeant for existing id positions********************************************************
# **********************************scan sergeant for existing id positions********************************************************


searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
# Click Edit Questions
editquestions = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.LINK_TEXT, 'Questions'))).click()

current_sergeant_question_id_list = []
chgpage = 1
id_match_scan = True
while id_match_scan is True:
    # Read in id text
    sergeant_question_id_list = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[starts-with(@class,'switch-container')]/div[3]/span/strong")))
    # Create secondary list obj
    all_my_sergeant_ids = [x.text for x in sergeant_question_id_list]
    for counter, myvar in enumerate(all_my_sergeant_ids):
        if myvar not in current_sergeant_question_id_list:
            current_sergeant_question_id_list.append(myvar)
            print(myvar)
        elif counter == len(all_my_sergeant_ids) - 1:
            try:
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/child::div[3]/div/ul/child::li[@class='next']/a"))).click()
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]")))
                chgpage += 1
            except Exception:
                print("Done retrieving all Sergeant ID's from page: " + str(chgpage))
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='survey-edit-questions']/div[3]/div/ul/li[1]/a"))).click()
                id_match_scan = False
                break
print(current_sergeant_question_id_list)

# handle conditions for when to delete or not delete.
# for i, excelrowlistsobject in enumerate(questionarr):
#      if excelrowlistobject[1] not in all_my_sergeant_ids:
#           continue
#       elif excelrowlistobject[2] == "Empty Slot":
#           continue
#       elif excelrowlistobject[1] is None and excelrowlistobject[2] is not None:
# delete stuff here
# elif condition to check for Y/N toggle for custom question where ID is present should be included
# above
