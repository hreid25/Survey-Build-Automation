import openpyxl
import time
import logging
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.chrome.options import Options
import re
from selenium.webdriver.common.by import By


# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************
# ***********************Start timer, load workbook, ask input********************************************************

start_time = time.time()
wb = openpyxl.load_workbook('Delete Edit Test QIL.xlsm')
# wb = openpyxl.load_workbook(input("Please enter the name of your QIL Document: "))
surveyhovers = wb['5- Hovers (Optional)']
surveyquest = wb['4- Survey Questions']
surveyinv = wb['2- Survey Invitation']

# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************
# ***********************CREATE MULTILINGUAL QIL ARRAY********************************************************

qarr = []
surveyquest.delete_cols(8)
for r in range(24, 177):
    questions = []
    for c in range(3, 101, 2):
        questions.append(surveyquest.cell(row=r, column=c).value)
    qarr.append(questions)
questionarr = [x for x in qarr if x != []]

for i, category in enumerate(questionarr):
    if category[0] is None:
        replacestring = str(questionarr[i - 1][0])
        questionarr[i][0] = replacestring


# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER************************************************************
# ****************************************************CREATE CHROME DRIVER*************************************************************

chrome_options = Options()
chrome_options.add_argument(
    "user-data-dir=C:\\Users\\haydr\\AppData\\Local\\Google\\Chrome\\User Data")
driver = webdriver.Chrome(chrome_options=chrome_options)

driver.get('https://www.infotech.com/kip')
driver.get('https://surveys.mcleanco.com/')

linkElem = driver.find_element_by_link_text('McLean & Company')
linkElem.click()
linkElemEng = driver.find_element_by_link_text('Engagement')
linkElemEng.click()

# surveyname = input("Enter the name of your survey as it appears in Sergeant: ")
surveyname = "Survey Automation Testing"

# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************
# **********************************GENERATE DRIVER SLUG MATCH********************************************************


searchforsurvey = driver.find_element_by_xpath(
    "//*[@id='q_translations_name_or_reseller_name_or_user_company_name_cont']")
searchforsurvey.send_keys(surveyname)
# click the search button after sending surveyname
driver.find_element_by_xpath('//*[@id="survey_search"]/input[4]').click()
findsurveys = driver.find_element_by_xpath("//tbody//a[text()='" + surveyname + "']").click()
driver.find_element_by_link_text("Edit Survey").click()
# click toggle
WebDriverWait(driver, 20).until(EC.element_to_be_clickable(
    (By.XPATH, "//*[@id='edit-tabs-dropdown']"))).click()
# Click Question Groups (Drivers) Option
WebDriverWait(driver, 20).until(EC.presence_of_element_located(
    (By.XPATH, "//*[@id='report-question-groups-nav']/a"))).click()

chgpage = 1
final_list = []
driver_match_scan = True
while driver_match_scan is True:
    # Read in table data
    all_the_drivers = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located(
        (By.XPATH, "//*[@id='report-question-groups-table']/tbody/tr")))
    # Create secondary list obj
    prettyname_slugname = [(x.text[0], x.text[1]) for x in all_the_drivers]
    print(prettyname_slugname)
    for counter, myvar in enumerate(prettyname_slugname):
        if myvar not in final_list:
            final_list.append(myvar)
        if counter == len(prettyname_slugname) - 1:
            try:
                # Click next
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@id='report-question-groups']/div[1]/div/ul/li[@class='next']/a"))).click()
                chgpage += 1
                # Checks to see which page is active
                WebDriverWait(driver, 5).until(EC.visibility_of_element_located(
                    (By.XPATH, "//*[contains(@class,'page active')]/a[@data-remote='true' and contains(text(),'" + str(chgpage) + "')]/ancestor::ul/preceding::tbody/tr")))
            except Exception as click_error:
                print("Done!...Returning to Page One", click_error)
                driver_match_scan = False
                # Return to page one
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                    (By.XPATH, "//*[@class='pagination']/li[@class='first']/a"))).click()
print(final_list)
# https://stackoverflow.com/questions/2397141/how-to-initialize-a-two-dimensional-array-in-python
# https://docs.python.org/3/tutorial/datastructures.html#dictionaries
# dropdown
# //*[@id = "edit-tabs-dropdown"]
# Select.option("Question Groups (Drivers)")
# print("--- %s seconds ---" % (time.time() - start_time))
# textarea for title, slug and question id
# //*[@id = "report-question-groups-table"]/tbody/tr

# //*[@id = "report-question-groups"]/div[1]/div/ul/li[4]
